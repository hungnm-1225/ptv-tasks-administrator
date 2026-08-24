# backend/app/services/cof_excel_service.py
import os
import logging
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Font

logger = logging.getLogger(__name__)


class COFExcelService:
    """Service chuyên trách bóc tách, chuẩn hóa dữ liệu COF và sinh file accounts.xlsx."""

    @staticmethod
    def _clean_str(val: Any) -> str:
        if val is None:
            return ""
        return str(val).strip()

    @staticmethod
    def _format_date_dob(dob_raw: Any) -> str:
        """Chuẩn hóa ngày sinh an toàn tuyệt đối về D/M/YYYY (Ví dụ: 1/1/1990 hoặc 15/8/2012)."""
        if not dob_raw:
            return "1/1/2000"

        from datetime import datetime, date
        if isinstance(dob_raw, (datetime, date)):
            return f"{dob_raw.day}/{dob_raw.month}/{dob_raw.year}"

        s = str(dob_raw).strip()
        s = s.split(" ")[0].split("T")[0]
        s = s.replace("-", "/").replace(".", "/")
        parts = [p.strip() for p in s.split("/") if p.strip()]

        if len(parts) == 3:
            try:
                if len(parts[0]) == 4:  # YYYY/MM/DD
                    return f"{int(parts[2])}/{int(parts[1])}/{int(parts[0])}"
                elif len(parts[2]) == 4:  # DD/MM/YYYY hoặc D/M/YYYY
                    return f"{int(parts[0])}/{int(parts[1])}/{int(parts[2])}"
            except ValueError:
                pass

        return "1/1/2000"

    @classmethod
    def parse_cof_file(cls, file_path: str) -> Dict[str, Any]:
        """Bóc tách toàn diện file COF: Tab 1 (Courses), Tab 2 (Students), Tab 3 (Teachers)."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Không tìm thấy file: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames

        # 1. BÓC TÁCH TAB 1: CURRICULUM ORDER FORM (COF)
        tab1_name = next((s for s in sheet_names if "curriculum" in s.lower() or "cof" in s.lower()), sheet_names[0])
        ws1 = wb[tab1_name]

        school_name = ""
        country = ""
        courses_to_order = []

        for row in range(1, 35):
            row_vals = [cls._clean_str(ws1.cell(row=row, column=c).value) for c in range(1, 15)]
            row_text = " ".join(row_vals).lower()
            
            if "school name:" in row_text:
                for idx, v in enumerate(row_vals):
                    if "school name:" in v.lower() and idx + 1 < len(row_vals):
                        school_name = row_vals[idx + 1] or (row_vals[idx + 2] if idx + 2 < len(row_vals) else "")
                        break
            if "country:" in row_text:
                for idx, v in enumerate(row_vals):
                    if "country:" in v.lower() and idx + 1 < len(row_vals):
                        country = row_vals[idx + 1]
                        break

        for r in range(25, ws1.max_row + 1):
            course_name = cls._clean_str(ws1.cell(row=r, column=4).value) or cls._clean_str(ws1.cell(row=r, column=3).value)
            category = cls._clean_str(ws1.cell(row=r, column=7).value) or "SWRP"
            course_id = cls._clean_str(ws1.cell(row=r, column=8).value)
            start_date = cls._clean_str(ws1.cell(row=r, column=9).value)
            end_date = cls._clean_str(ws1.cell(row=r, column=12).value)
            qty_val = cls._clean_str(ws1.cell(row=r, column=17).value)

            if not start_date or not end_date or start_date == "---" or end_date == "---":
                continue
            if "dd-mm-yyyy" in start_date.lower() or "course name" in course_name.lower() or "school start date" in start_date.lower():
                continue

            try:
                licenses = int(float(qty_val)) if qty_val else 1
            except ValueError:
                licenses = 1

            courses_to_order.append({
                "course_id": str(course_id).replace(".0", ""),
                "course_name": course_name,
                "category": category,
                "start_date": start_date,
                "end_date": end_date,
                "licenses": licenses,
                "row_index": r
            })

        # 2. BÓC TÁCH TAB 2: STUDENT INFORMATION
        tab2_name = next((s for s in sheet_names if "student" in s.lower()), None)
        students_all = []
        students_to_create = []

        if tab2_name:
            ws2 = wb[tab2_name]
            for r in range(7, ws2.max_row + 1):
                full_name = cls._clean_str(ws2.cell(row=r, column=2).value)
                first_name = cls._clean_str(ws2.cell(row=r, column=3).value)
                last_name = cls._clean_str(ws2.cell(row=r, column=4).value)
                email = cls._clean_str(ws2.cell(row=r, column=6).value)
                dob = cls._clean_str(ws2.cell(row=r, column=7).value)
                account_exist = cls._clean_str(ws2.cell(row=r, column=8).value).lower()
                grade = cls._clean_str(ws2.cell(row=r, column=9).value)
                class_group = cls._clean_str(ws2.cell(row=r, column=11).value)
                username = cls._clean_str(ws2.cell(row=r, column=12).value)

                if not first_name and not full_name:
                    continue
                if "total" in full_name.lower() or "total" in first_name.lower():
                    continue

                student_record = {
                    "row_index": r,
                    "full_name": full_name or f"{first_name} {last_name}".strip(),
                    "first_name": first_name or (full_name.split()[0] if full_name else "Student"),
                    "last_name": last_name or (" ".join(full_name.split()[1:]) if len(full_name.split()) > 1 else "Auto"),
                    "email": email,
                    "dob": cls._format_date_dob(dob),
                    "grade": grade,
                    "class_group": class_group or "Default Class",
                    "username": username,
                    "role": "student"
                }
                students_all.append(student_record)

                if not username and account_exist != "yes":
                    students_to_create.append(student_record)

        # 3. BÓC TÁCH TAB 3: TEACHER INFORMATION
        tab3_name = next((s for s in sheet_names if "teacher" in s.lower()), None)
        teachers_all = []
        teachers_to_create = []

        if tab3_name:
            ws3 = wb[tab3_name]
            for r in range(7, ws3.max_row + 1):
                class_group = cls._clean_str(ws3.cell(row=r, column=3).value)
                full_name = cls._clean_str(ws3.cell(row=r, column=5).value)
                first_name = cls._clean_str(ws3.cell(row=r, column=6).value)
                last_name = cls._clean_str(ws3.cell(row=r, column=7).value)
                email = cls._clean_str(ws3.cell(row=r, column=8).value)
                dob = cls._clean_str(ws3.cell(row=r, column=9).value)
                account_exist = cls._clean_str(ws3.cell(row=r, column=10).value).lower()
                course_assign = cls._clean_str(ws3.cell(row=r, column=11).value)
                username = cls._clean_str(ws3.cell(row=r, column=12).value)

                if not first_name and not full_name:
                    continue
                if "total" in full_name.lower() or "total" in class_group.lower() or "total" in first_name.lower():
                    continue

                teacher_record = {
                    "row_index": r,
                    "full_name": full_name or f"{first_name} {last_name}".strip(),
                    "first_name": first_name or (full_name.split()[0] if full_name else "Teacher"),
                    "last_name": last_name or "Auto",
                    "email": email,
                    "dob": cls._format_date_dob(dob),
                    "class_group": class_group or "Default Class",
                    "course_assign": course_assign,
                    "username": username,
                    "role": "teacher"
                }
                teachers_all.append(teacher_record)

                if not username and account_exist != "yes":
                    teachers_to_create.append(teacher_record)

        wb.close()

        return {
            "school_name": school_name,
            "country": country,
            "courses": courses_to_order,
            "students_all": students_all,
            "students_to_create": students_to_create,
            "teachers_all": teachers_all,
            "teachers_to_create": teachers_to_create
        }

    @classmethod
    def generate_accounts_excel(cls, accounts_to_create: List[Dict[str, Any]], output_path: str) -> int:
        """Sinh file accounts.xlsx chuẩn Template bắt đầu từ dòng 6."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Class 7s"

        ws.cell(row=2, column=2, value="Account creation request form")

        headers = [
            "No.", 
            "First Name (*)", 
            "Last Name (*)", 
            "Mobile number (Optional)", 
            "Email (*)", 
            "Date of Birth (*)(DD/MM/YYYY)", 
            "Role (*)"
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=5, column=col_num, value=header)

        count = 0
        current_row = 6
        for idx, acc in enumerate(accounts_to_create, 1):
            role_formatted = acc.get("role", "Student").capitalize()
            dob_formatted = cls._format_date_dob(acc.get("dob"))
            email_val = acc.get("email") or ""

            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=acc.get("first_name", ""))
            ws.cell(row=current_row, column=3, value=acc.get("last_name", ""))
            ws.cell(row=current_row, column=4, value="")
            ws.cell(row=current_row, column=5, value=email_val)
            ws.cell(row=current_row, column=6, value=dob_formatted)
            ws.cell(row=current_row, column=7, value=role_formatted)

            count += 1
            current_row += 1

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        wb.close()
        logger.info(f"💾 Đã tạo file Template Form ({count} tài khoản) tại: {output_path}")
        return count

    @classmethod
    def detect_and_process_excel(cls, file_path: str, temp_dir: str) -> Tuple[str, int, int, int, bool, Dict[str, Any]]:
        """
        Tự động nhận diện file tải lên là:
        1. File COF 3 Tabs -> Bóc tách sinh accounts.xlsx
        2. File Accounts 7 cột -> Lấy dùng trực tiếp
        Trả về: (path_accounts_file, student_count, teacher_count, total_count, is_cof_file, cof_parsed_data)
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_lower = [s.lower() for s in wb.sheetnames]
        is_cof = any("student" in s or "curriculum" in s or "cof" in s for s in sheets_lower)

        if is_cof:
            wb.close()
            parsed = cls.parse_cof_file(file_path)
            students = parsed.get("students_to_create", [])
            teachers = parsed.get("teachers_to_create", [])
            all_accounts = students + teachers
            
            output_acc_path = os.path.join(temp_dir, f"ready_accounts_{os.path.basename(file_path)}")
            total_count = cls.generate_accounts_excel(all_accounts, output_acc_path)
            return output_acc_path, len(students), len(teachers), total_count, True, parsed
        else:
            # File dạng accounts.xlsx trực tiếp
            ws = wb.active
            students_c = 0
            teachers_c = 0
            for r in range(6, ws.max_row + 1):
                fn = cls._clean_str(ws.cell(row=r, column=2).value)
                role = cls._clean_str(ws.cell(row=r, column=7).value).lower()
                if not fn:
                    continue
                if "teacher" in role or "gv" in role or "giáo viên" in role:
                    teachers_c += 1
                else:
                    students_c += 1
            wb.close()
            total_c = students_c + teachers_c
            return file_path, students_c, teachers_c, total_c, False, {}

    @classmethod
    def write_results_back_to_cof(
        cls, 
        original_cof_path: str, 
        result_excel_path: Optional[str], 
        students_all: List[Dict[str, Any]], 
        students_to_create: List[Dict[str, Any]], 
        teachers_all: List[Dict[str, Any]], 
        teachers_to_create: List[Dict[str, Any]], 
        output_cof_path: str
    ) -> str:
        """Ghi ngược kết quả User/Pass và Group Name in LMS vào file COF."""
        wb_orig = openpyxl.load_workbook(original_cof_path)

        res_map = {}
        if result_excel_path and os.path.exists(result_excel_path):
            wb_res = openpyxl.load_workbook(result_excel_path, data_only=True)
            ws_res = wb_res.active
            for r in range(2, ws_res.max_row + 1):
                fn = cls._clean_str(ws_res.cell(row=r, column=1).value)
                ln = cls._clean_str(ws_res.cell(row=r, column=2).value)
                u = cls._clean_str(ws_res.cell(row=r, column=3).value)
                p = cls._clean_str(ws_res.cell(row=r, column=4).value)
                e = cls._clean_str(ws_res.cell(row=r, column=5).value)
                
                key = e.lower() if e else f"{fn}_{ln}".lower()
                if u:
                    res_map[key] = {"username": u, "password": p}
            wb_res.close()

        highlight_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        highlight_font = Font(name="Calibri", size=11, bold=True, color="C00000")

        # 2. Student Tab
        sheet_names = wb_orig.sheetnames
        tab2_name = next((s for s in sheet_names if "student" in s.lower()), None)
        if tab2_name:
            ws_student = wb_orig[tab2_name]
            create_rows = {acc["row_index"]: acc for acc in students_to_create}

            for idx, acc in enumerate(students_all):
                r = acc["row_index"]
                grp_name = acc.get("class_group", "")

                if grp_name:
                    cell_grp = ws_student.cell(row=r, column=14, value=grp_name)
                    cell_grp.fill = highlight_fill
                    cell_grp.font = highlight_font

                if r in create_rows:
                    key = acc["email"].lower() if acc["email"] else f"{acc['first_name']}_{acc['last_name']}".lower()
                    res = res_map.get(key)
                    if not res and idx < len(res_map):
                        res = list(res_map.values())[idx]

                    if res:
                        cell_u = ws_student.cell(row=r, column=12, value=res["username"])
                        cell_u.fill = highlight_fill
                        cell_u.font = highlight_font

                        cell_p = ws_student.cell(row=r, column=13, value=res["password"])
                        cell_p.fill = highlight_fill
                        cell_p.font = highlight_font

        # 3. Teacher Tab
        tab3_name = next((s for s in sheet_names if "teacher" in s.lower()), None)
        if tab3_name:
            ws_teacher = wb_orig[tab3_name]
            create_teacher_rows = {acc["row_index"]: acc for acc in teachers_to_create}

            for idx, acc in enumerate(teachers_all):
                r = acc["row_index"]
                grp_name = acc.get("class_group", "")

                if grp_name:
                    cell_grp = ws_teacher.cell(row=r, column=14, value=grp_name)
                    cell_grp.fill = highlight_fill
                    cell_grp.font = highlight_font

                if r in create_teacher_rows:
                    key = acc["email"].lower() if acc["email"] else f"{acc['first_name']}_{acc['last_name']}".lower()
                    res = res_map.get(key)
                    if not res and idx < len(res_map):
                        res = list(res_map.values())[idx]

                    if res:
                        cell_u = ws_teacher.cell(row=r, column=12, value=res["username"])
                        cell_u.fill = highlight_fill
                        cell_u.font = highlight_font

                        cell_p = ws_teacher.cell(row=r, column=13, value=res["password"])
                        cell_p.fill = highlight_fill
                        cell_p.font = highlight_font

        os.makedirs(os.path.dirname(output_cof_path), exist_ok=True)
        wb_orig.save(output_cof_path)
        wb_orig.close()
        logger.info(f"✨ File COF hoàn thiện đã lưu tại: {output_cof_path}")
        return output_cof_path